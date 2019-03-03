#!/usr/bin/env python3

import openpyxl
import calendar
import datetime

year = 2019

sheet_names = [
        'Januar',
        'Februar',
        'Mars',
        'April',
        'Mai',
        'Juni',
        'Juli',
        'August',
        'Septemeber',
        'Oktober',
        'November',
        'Desember'
]

days = [
        'mandag',
        'tirsdag',
        'onsdag',
        'torsdag',
        'fredag',
        'lørdag',
        'søndag'
]


wb = openpyxl.Workbook()

for i, s in enumerate(sheet_names):
    wb.create_sheet(title=s, index=i)


for month, s in enumerate(sheet_names, start=1):
    ws = wb[s]
    ws['A1'] = 'Uke nr'
    ws['B1'] = 'Dag'
    ws['C1'] = 'Dato'
    ws['D1'] = 'Kom (tidspunkt)'
    ws['E1'] = 'Gikk (tidspunkt)'
    ws['F1'] = 'Antall timer'
    ws['G1'] = 'Antall timer denne uken'
    
    prev_week_nr = 1
    week_start_it = 2
    it = 2
    print('Month = ', s, month)
    for day in range(1, calendar.monthlen(year, 1)+1):

        
        print('Day = ', day)
        try:
            date = datetime.date(year, month, day)
        except ValueError:
            continue
        week_nr = date.isocalendar()[1]
        

        if week_nr != prev_week_nr:
            week_start_it = it
            prev_week_nr = week_nr
            
        
        ws[f'A{it}'] = week_nr
        ws[f'B{it}'] = days[date.weekday()]
        ws[f'C{it}'] = date.isoformat()
        ws[f'F{it}'] = f'=E{it}-D{it}'
        ws[f'G{it}'] = f'=SUM(F{week_start_it}:F{it})'
        
        it += 1

    ws[f'A{it+1}'] = 'Totalt anntall timer denne måneden:'
    ws[f'E{it+1}'] = f'=SUM(F2:F{it})'
        
wb.save('test.xlsx')
